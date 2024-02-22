import openpyxl
from openpyxl.styles import PatternFill
from matplotlib.colors import Normalize, to_rgba, ListedColormap, colorConverter
import os
import pandas as pd
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.cm as cm
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import numpy as np


# 桁数を変換
def convert_value(value):
    if (isinstance(value, float)) or (pd.isna(value)):
        return value
    elif 'k' in value:
        return float(value.replace('k', '')) * 1000
    elif 'M' in value:
        return float(value.replace('M', '')) * 1000000
    elif 'B' in value:
        return float(value.replace('B', '')) * 1000000000
    else:
        return None
# RGBの配列を16進数に変換
def rgb_to_hex(rgb):
    hex_color = [round(x * 255) for x in rgb[:3]]
    hex_color = [format(x, '02X') for x in hex_color]
    hex_code = ''.join(hex_color)
    return hex_code
# 与えられた日付の配列から1ヶ月以内の日付を判別
def is_within_one_month(date_str):
    # 月の辞書を定義
    month_dict = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}

    # 今日の日付を取得
    now = datetime.now().strftime("%m-%d")
    # 1ヶ月後の日付を取得
    next_one_month = (datetime.now() + timedelta(days=30)).strftime("%m-%d")
    # 与えられた月の値
    month = month_dict.get(date_str[:3])
    # 与えられた日の値
    date = int(date_str[4:6])
    # 与えられた日付
    month_date = datetime(datetime.now().year, month, date).strftime("%m-%d")
    return (now <= month_date) and (month_date <= next_one_month)
# ヒートマップ作成の初期実行
def init_heatmap(column_name, cmap, vmax, vmin):
    df[column_name] = pd.to_numeric(df[column_name], errors="coerce")
    value = df[[column_name]].values
    cmap = cm.get_cmap(cmap)
    data = value.flatten()
    norm = Normalize(vmax=vmax, vmin=vmin)
    rgba_values = cmap(norm(data))
    hex_colors = [rgb_to_hex(rgb) for rgb in rgba_values]
    return hex_colors


# csvファイルをxlsxファイルに変換
df_dir = os.getcwd()+"/Stock_Trade/BuyingStock.csv"
df = pd.read_csv(df_dir)
df.to_excel(os.getcwd()+"/Stock_Trade/BuyingStock.xlsx", encoding='utf-8')
df_xlsx_dir = os.getcwd()+"/Stock_Trade/BuyingStock.xlsx"

# xlsxファイルを読み込み
wb = openpyxl.load_workbook(df_xlsx_dir)
ws = wb['Sheet1']

# InsiderOwn
for i in range(ws.max_row-1):
    cell_market_cap = ws.cell(row=i+2, column=6)
    cell_insider_own = ws.cell(row=i+2, column=9)
    market_cap = convert_value(cell_market_cap.value)
    insider_own = float(cell_insider_own.value.replace("%", ""))
    if market_cap > 10e+9: # 大型株を判別
        if insider_own >= 1: # 該当している場合は背景を緑色に変更
            fill_color = PatternFill(fgColor='008000', bgColor='008000', fill_type='solid')
            cell_insider_own.fill = fill_color
    else: # その他は中小型株
        if insider_own >= 3: # 該当している場合は背景を緑色に変更
            fill_color = PatternFill(fgColor='008000', bgColor='008000', fill_type='solid')
            cell_insider_own.fill = fill_color

# EarningsDate
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=10)
    earnings_date = cell.value
    if is_within_one_month(earnings_date):
        fill_color = PatternFill(fgColor='ffff00', bgColor='ffff00', fill_type='solid')
        cell.fill = fill_color

# RS
hex_colors = init_heatmap(column_name='RS', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=13)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color

# ATH
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_ath = ws.cell(row=i+2, column=15)
    price = format(float(cell_price.value), '.2f')
    ath = format(float(cell_ath.value), '.2f')
    if ath == price:
        fill_color = PatternFill(fgColor='008000', bgColor='008000', fill_type='solid')
        cell_ath.fill = fill_color

# 52W High
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_w52high = ws.cell(row=i+2, column=16)
    price = format(float(cell_price.value), '.2f')
    w52high = format(float(cell_w52high.value), '.2f')
    if w52high == price:
        fill_color = PatternFill(fgColor='008000', bgColor='008000', fill_type='solid')
        cell_w52high.fill = fill_color

# SMA200 Gap
hex_colors = init_heatmap(column_name='SMA200 Gap', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=31)
    sma200_gap = float(format(cell.value, '.2f'))
    if sma200_gap >= 1.1: # 10%の閾値は検討する
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell.fill = fill_color

# Volume SMA50
hex_colors = init_heatmap(column_name='Volume SMA50', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_volume = ws.cell(row=i+2, column=11)
    cell_volume_avg = ws.cell(row=i+2, column=32)
    volume = float(cell_volume.value.replace(",", ""))
    volume_avg = float(cell_volume_avg.value)
    if volume >= volume_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_volume.fill = fill_color

# SMA10
hex_colors = init_heatmap(column_name='SMA10', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_price_avg = ws.cell(row=i+2, column=25)
    price = float(cell_price.value)
    price_avg = float(format(cell_price_avg.value, '.2f'))
    if price >= price_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_price_avg.fill = fill_color

# EMA8
hex_colors = init_heatmap(column_name='EMA8', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_price_avg = ws.cell(row=i+2, column=26)
    price = float(cell_price.value)
    price_avg = float(format(cell_price_avg.value, '.2f'))
    if price >= price_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_price_avg.fill = fill_color

# EMA21
hex_colors = init_heatmap(column_name='EMA21', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_price_avg = ws.cell(row=i+2, column=27)
    price = float(cell_price.value)
    price_avg = float(format(cell_price_avg.value, '.2f'))
    if price >= price_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_price_avg.fill = fill_color

# SMA50
hex_colors = init_heatmap(column_name='SMA50', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_price_avg = ws.cell(row=i+2, column=28)
    price = float(cell_price.value)
    price_avg = float(format(cell_price_avg.value, '.2f'))
    if price >= price_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_price_avg.fill = fill_color

# SMA200
hex_colors = init_heatmap(column_name='SMA200', cmap='YlGn', vmax=None, vmin=None)
for i in range(ws.max_row-1):
    cell_price = ws.cell(row=i+2, column=12)
    cell_price_avg = ws.cell(row=i+2, column=29)
    price = float(cell_price.value)
    price_avg = float(format(cell_price_avg.value, '.2f'))
    if price >= price_avg:
        fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
        cell_price_avg.fill = fill_color

# Previous Quarter Eps2
# hex_colors = init_heatmap(column_name='Previous Quarter EPS2', cmap='RdYlGn', vmax=50, vmin=-50)
hex_colors = init_heatmap(column_name='Previous Quarter Eps2', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=37)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Quarter EPS
hex_colors = init_heatmap(column_name='Previous Quarter EPS', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=38)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Corrent Quarter EPS
hex_colors = init_heatmap(column_name='Corrent Quarter EPS', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=39)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Next Quarter EPS
hex_colors = init_heatmap(column_name='Next Quarter EPS', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=40)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Annual EPS2
hex_colors = init_heatmap(column_name='Previous Annual EPS2', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=41)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Annual EPS
hex_colors = init_heatmap(column_name='Previous Annual EPS', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=42)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Next Annual EPS
hex_colors = init_heatmap(column_name='Next Annual EPS', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=43)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Quarter Revenue2
hex_colors = init_heatmap(column_name='Previous Quarter Revenue2', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=44)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Quarter Revenue
hex_colors = init_heatmap(column_name='Previous Quarter Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=45)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Corrent Quarter Revenue
hex_colors = init_heatmap(column_name='Corrent Quarter Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=46)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Next Quarter Revenue
hex_colors = init_heatmap(column_name='Next Quarter Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=47)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Annual Revenue2
hex_colors = init_heatmap(column_name='Previous Annual Revenue2', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=48)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Previous Annual Revenue
hex_colors = init_heatmap(column_name='Previous Annual Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=49)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Corrent Annual Revenue
hex_colors = init_heatmap(column_name='Corrent Annual Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=50)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color
# Next Annual Revenue
hex_colors = init_heatmap(column_name='Next Annual Revenue', cmap='RdYlGn', vmax=50, vmin=-50)
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=51)
    fill_color = PatternFill(fgColor=hex_colors[i], bgColor=hex_colors[i], fill_type='solid')
    cell.fill = fill_color

# Perfome Institute
for i in range(ws.max_row-1):
    cell = ws.cell(row=i+2, column=52)
    if cell.value!=None:
        perf_inst = float(cell.value.replace('%', ''))
        if perf_inst >= 0:
            fill_color = PatternFill(fgColor='008000', bgColor='008000', fill_type='solid')
            cell.fill = fill_color
        else:
            fill_color = PatternFill(fgColor='a80000', bgColor='ff0000', fill_type='solid')
            cell.fill = fill_color

# 不要な列を削除する
# delete_col = [6, 7, 13, 17, 20, 21, 22, 23, 32, 33, 34]
# for i in range(len(delete_col)):
#     #列を削除
#     ws.delete_cols(delete_col[i])

# xlsxファイルの保存
wb.save(os.getcwd()+"/Stock_Trade/BuyingStock.xlsx")
